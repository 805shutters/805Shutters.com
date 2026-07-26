#!/usr/bin/env python3
"""Generate the auditable Norman Roller V2 catalog and limit source.

The workbook is authoritative. This generator expands merged headers, keeps
the source row/cell lineage, and only creates selectable profiles when every
required width/height limit is present in inches. Source defects are retained
and fail closed rather than being guessed or silently corrected.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import OrderedDict, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


EXPECTED_INGESTED_SHA256 = "f076f92a2f9f5032c78c48487afb86464b8197d567822efd7d8dbb79dd18e253"
EXPECTED_ORIGINAL_SHA256 = "ba286767b6c4d760bf480434312678f6c35b229e9763ddd6fcda9f7cd18b9cc3"
INGESTION_AS_OF = date(2026, 7, 20)

LIMIT_SHEETS = (
    "Single(Non-LG360)&Common",
    "LG360&w T-post split & housing",
    "LG360 with T-Post (2 ) (Std)",
    "LG360 with T-Post (2 ) (Ind)",
    "LG360 with T-Post (3 Shades)",
    "LG360 with T-Post (4 Shades)",
    "Standard Coupled Shade(2)",
    "Independently Coupled Shade(2)",
    "Dual",
    "Cassette",
    "Coupled Shades(3)",
    "Coupled Shades(4)",
)

EXPECTED_COUNTS = {
    "collections": 73,
    "colors": 350,
    "offerings": 373,
    "fabricCodes": 89,
    "dimensionSheets": 12,
    "dimensionColumns": 594,
    "rawLimitRows": 937,
    "quarantinedLimitRows": 1,
    "activeLimitRows": 936,
    "rawNumericCells": 46372,
    "quarantinedNumericCells": 40,
    "activeNumericCells": 46332,
    "profileDefinitions": 149,
    "usableProfileDefinitions": 144,
    "unusableProfileDefinitions": 5,
}

EXPECTED_REGION_COUNTS = {
    "all_regions": 327,
    "other_regions": 23,
    "ca_ma": 23,
}

FABRIC_TOKEN = re.compile(r"(?:[A-Z]\d{5}|[A-Z]{2}\d{4}(?:-[A-Z])?)")
REQUIRED_LIMIT_METRICS = ("minWidth", "minHeight", "maxWidth", "maxHeight")

METRIC_NAMES = {
    "min width": "minWidth",
    "min height": "minHeight",
    "max width": "maxWidth",
    "max height": "maxHeight",
    "max area": "maxAreaSqft",
    "total max area for two shades": "totalMaxAreaTwoShadesSqft",
    "max area for each shade": "maxAreaEachShadeSqft",
    "total max area for two shades of coupled shade": "totalMaxAreaCoupledPairSqft",
    "max area for single shade": "maxAreaSingleShadeSqft",
    "total max area for two shades of one coupled shade": "totalMaxAreaOneCoupledPairSqft",
}

UNIT_NAMES = {
    "inch": "inch",
    "sqft": "sqft",
    "mm": "mm",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u3000", " ").split())


def exact_text(value: object) -> str:
    return "" if value is None else str(value)


def numeric(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def json_number(value: int | float) -> int | float:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def source_original_for(converted: Path, explicit: Path | None) -> Path:
    if explicit is not None:
        return explicit
    candidates = (
        converted.with_suffix(".xls"),
        converted.parent.parent / f"{converted.stem}.xls",
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise SystemExit(
        "Original .xls source not found. Pass --original-workbook so both source identities are pinned."
    )


class ExpandedSheet:
    """Read merged cells as if each cell contained its top-left value."""

    def __init__(self, sheet: Any):
        self.sheet = sheet
        self._merged: dict[tuple[int, int], tuple[object, int, int]] = {}
        for merged_range in sheet.merged_cells.ranges:
            value = sheet.cell(merged_range.min_row, merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for column in range(merged_range.min_col, merged_range.max_col + 1):
                    self._merged[(row, column)] = (
                        value,
                        merged_range.min_row,
                        merged_range.min_col,
                    )

    def value(self, row: int, column: int) -> object:
        merged = self._merged.get((row, column))
        return merged[0] if merged else self.sheet.cell(row, column).value

    def anchor(self, row: int, column: int) -> str:
        merged = self._merged.get((row, column))
        if merged:
            row, column = merged[1], merged[2]
        return f"{get_column_letter(column)}{row}"


def normalize_region(source_note: str) -> str:
    note = clean(source_note)
    if not note:
        return "all_regions"
    if note == "For all other regions":
        return "other_regions"
    if note == "For CA/MA states only":
        return "ca_ma"
    raise SystemExit(f"Unknown regional scope in Fabric Code List: {note!r}")


def source_release(workbook: Any) -> dict[str, Any]:
    sheet = workbook["Revision Log"]
    dated_rows: list[tuple[date, date, int]] = []
    for row in range(1, sheet.max_row + 1):
        revised = sheet.cell(row, 1).value
        effective = sheet.cell(row, 2).value
        if isinstance(revised, (date, datetime)) and isinstance(effective, (date, datetime)):
            revised_date = revised.date() if isinstance(revised, datetime) else revised
            effective_date = effective.date() if isinstance(effective, datetime) else effective
            dated_rows.append((revised_date, effective_date, row))
    if not dated_rows:
        raise SystemExit("Revision Log has no dated release row")
    revised, effective, row = max(dated_rows, key=lambda entry: (entry[1], entry[0], entry[2]))
    status = "future" if INGESTION_AS_OF < effective else "active"
    return {
        "revisionDate": revised.isoformat(),
        "effectiveFrom": effective.isoformat(),
        "releaseStatus": status,
        "releaseStatusEvaluatedOn": INGESTION_AS_OF.isoformat(),
        "activationPolicy": "inactive before effectiveFrom; active on or after effectiveFrom",
        "sourceSummary": exact_text(sheet.cell(row, 3).value),
        "sourceRef": {
            "sheet": "Revision Log",
            "row": row,
            "range": f"A{row}:C{row}",
        },
    }


def parse_fabric_catalog(workbook: Any, effective_from: str) -> tuple[list[dict], list[dict], list[str]]:
    sheet = workbook["Fabric Code List"]
    expanded = ExpandedSheet(sheet)
    offerings: list[dict] = []

    for row in range(2, sheet.max_row + 1):
        fabric_raw = sheet.cell(row, 2).value
        color_raw = sheet.cell(row, 4).value
        if not clean(fabric_raw) or not clean(color_raw):
            continue
        source_note_raw = expanded.value(row, 6)
        offerings.append(
            {
                "id": f"roller-offering-{len(offerings) + 1:03d}",
                "itemNumber": json_number(sheet.cell(row, 1).value)
                if numeric(sheet.cell(row, 1).value)
                else None,
                "fabricCode": clean(fabric_raw).upper(),
                "sourceFabricCode": exact_text(fabric_raw),
                "collection": clean(sheet.cell(row, 3).value),
                "sourceCollection": exact_text(sheet.cell(row, 3).value),
                "colorCode": clean(color_raw).upper(),
                "sourceColorCode": exact_text(color_raw),
                "colorName": clean(sheet.cell(row, 5).value),
                "sourceColorName": exact_text(sheet.cell(row, 5).value),
                "regionScope": normalize_region(exact_text(source_note_raw)),
                "sourceRegionNote": exact_text(source_note_raw),
                "effectiveFrom": effective_from,
                "sourceRef": {
                    "sheet": "Fabric Code List",
                    "row": row,
                    "range": f"B{row}:F{row}",
                    "fabricCodeCell": f"B{row}",
                    "collectionCell": f"C{row}",
                    "colorCodeCell": f"D{row}",
                    "colorNameCell": f"E{row}",
                    "regionCell": expanded.anchor(row, 6),
                },
            }
        )

    color_map: OrderedDict[str, dict] = OrderedDict()
    for offering in offerings:
        code = offering["colorCode"]
        identity = (offering["colorName"], offering["collection"])
        if code in color_map:
            existing_identity = (color_map[code]["colorName"], color_map[code]["collection"])
            if existing_identity != identity:
                raise SystemExit(
                    f"Conflicting source identity for color {code}: {existing_identity!r} vs {identity!r}"
                )
            color_map[code]["offeringIds"].append(offering["id"])
            if offering["fabricCode"] not in color_map[code]["fabricCodes"]:
                color_map[code]["fabricCodes"].append(offering["fabricCode"])
            if offering["regionScope"] not in color_map[code]["regionScopes"]:
                color_map[code]["regionScopes"].append(offering["regionScope"])
        else:
            color_map[code] = {
                "colorCode": code,
                "colorName": offering["colorName"],
                "collection": offering["collection"],
                "fabricCodes": [offering["fabricCode"]],
                "regionScopes": [offering["regionScope"]],
                "effectiveFrom": effective_from,
                "offeringIds": [offering["id"]],
            }

    collections = list(OrderedDict.fromkeys(row["collection"] for row in offerings))
    return offerings, list(color_map.values()), collections


def metric_for(header_path: Iterable[str]) -> tuple[str | None, str | None]:
    for source_value in reversed(tuple(header_path)):
        normalized = clean(source_value).lower()
        if normalized in METRIC_NAMES:
            return METRIC_NAMES[normalized], clean(source_value)
    return None, None


def unit_for(header_path: Iterable[str]) -> tuple[str | None, str | None]:
    for source_value in reversed(tuple(header_path)):
        normalized = clean(source_value).lower()
        if normalized in UNIT_NAMES:
            return UNIT_NAMES[normalized], clean(source_value)
    return None, None


def tube_for(header_path: Iterable[str]) -> str | None:
    for source_value in header_path:
        if "tube" in clean(source_value).lower():
            return clean(source_value)
    return None


def orientation_for(header_path: Iterable[str]) -> str | None:
    for source_value in header_path:
        if "orientation" in clean(source_value).lower():
            return clean(source_value)
    return None


def application_for(header_path: Iterable[str]) -> str | None:
    for source_value in header_path:
        normalized = clean(source_value).lower()
        # The sheet title can itself contain "valance". The actual application
        # header in this source explicitly names top treatment and must win.
        if "top treatment" in normalized:
            return clean(source_value)
    return None


def column_record(
    expanded: ExpandedSheet,
    sheet_index: int,
    column: int,
) -> dict[str, Any]:
    source_header_cells: list[dict[str, Any]] = []
    seen_anchors: set[str] = set()
    header_path: list[str] = []
    for row in range(1, 8):
        value = expanded.value(row, column)
        if not clean(value):
            continue
        anchor = expanded.anchor(row, column)
        if anchor not in seen_anchors:
            source_header_cells.append(
                {
                    "sourceCell": anchor,
                    "sourceRow": row,
                    "value": exact_text(value),
                }
            )
            seen_anchors.add(anchor)
        cleaned = clean(value)
        if not header_path or header_path[-1] != cleaned:
            header_path.append(cleaned)

    metric, source_metric = metric_for(header_path)
    unit, source_unit = unit_for(header_path)
    operating_system = clean(expanded.value(3, column)) or None
    letter = get_column_letter(column)
    return {
        "id": f"roller-sheet-{sheet_index:02d}-col-{letter.lower()}",
        "sourceColumn": letter,
        "sourceColumnIndex": column,
        "headerPath": header_path,
        "sourceHeaderCells": source_header_cells,
        "orientation": orientation_for(header_path),
        "operatingSystem": operating_system,
        "application": application_for(header_path),
        "tube": tube_for(header_path),
        "metric": metric,
        "sourceMetric": source_metric,
        "unit": unit,
        "sourceUnit": source_unit,
    }


def candidate_data_rows(expanded: ExpandedSheet) -> list[int]:
    sheet = expanded.sheet
    rows: list[int] = []
    for row in range(1, sheet.max_row + 1):
        # Category dividers merge A:C (for example "Room Darkening") and can
        # contain a stray zero in a limit column. Only a value physically
        # authored in column B establishes a fabric row. Column A remains
        # optional because the source has three valid, unnumbered fabric rows.
        if not clean(sheet.cell(row, 2).value):
            continue
        if any(numeric(sheet.cell(row, column).value) for column in range(4, sheet.max_column + 1)):
            rows.append(row)
    return rows


def extract_fabric_tokens(source_label: str) -> list[str]:
    return list(OrderedDict.fromkeys(FABRIC_TOKEN.findall(source_label.upper())))


def parse_limit_sheets(workbook: Any, known_fabric_codes: set[str]) -> tuple[list, list, list]:
    dimension_sheets: list[dict] = []
    active_rows: list[dict] = []
    quarantined_rows: list[dict] = []

    for sheet_index, sheet_name in enumerate(LIMIT_SHEETS, start=1):
        sheet = workbook[sheet_name]
        expanded = ExpandedSheet(sheet)
        data_rows = candidate_data_rows(expanded)
        numeric_columns = [
            column
            for column in range(4, sheet.max_column + 1)
            if any(numeric(sheet.cell(row, column).value) for row in data_rows)
        ]
        if not numeric_columns:
            raise SystemExit(f"No numeric restriction columns found in {sheet_name}")
        expected_contiguous = list(range(min(numeric_columns), max(numeric_columns) + 1))
        if numeric_columns != expected_contiguous:
            raise SystemExit(f"Non-contiguous restriction columns in {sheet_name}")

        columns = [column_record(expanded, sheet_index, column) for column in numeric_columns]
        unknown_headers = [column for column in columns if column["metric"] is None or column["unit"] is None]
        if unknown_headers:
            refs = ", ".join(column["sourceColumn"] for column in unknown_headers)
            raise SystemExit(f"Unrecognized metric/unit headers in {sheet_name}: {refs}")

        sheet_raw_count = 0
        sheet_active_count = 0
        sheet_quarantine_count = 0
        sheet_raw_cells = 0
        sheet_active_cells = 0
        sheet_quarantine_cells = 0

        for row in data_rows:
            source_fabric_label = clean(sheet.cell(row, 2).value).upper()
            source_tokens = extract_fabric_tokens(source_fabric_label)
            matched_codes = [token for token in source_tokens if token in known_fabric_codes]
            values = [
                json_number(sheet.cell(row, column).value)
                if numeric(sheet.cell(row, column).value)
                else None
                for column in numeric_columns
            ]
            numeric_cell_count = sum(value is not None for value in values)
            record = {
                "id": f"roller-sheet-{sheet_index:02d}-row-{row}",
                "sheet": sheet_name,
                "sourceRow": row,
                "sourceCellRange": (
                    f"{get_column_letter(numeric_columns[0])}{row}:"
                    f"{get_column_letter(numeric_columns[-1])}{row}"
                ),
                "itemNumber": json_number(expanded.value(row, 1))
                if numeric(expanded.value(row, 1))
                else None,
                "sourceItemCell": expanded.anchor(row, 1),
                "sourceFabricLabel": source_fabric_label,
                "sourceFabricCell": f"B{row}",
                "sourceFabricTokens": source_tokens,
                "fabricCodes": matched_codes,
                "collection": clean(expanded.value(row, 3)),
                "sourceCollection": exact_text(expanded.value(row, 3)),
                "sourceCollectionCell": expanded.anchor(row, 3),
                "values": values,
                "numericCellCount": numeric_cell_count,
            }

            sheet_raw_count += 1
            sheet_raw_cells += numeric_cell_count
            if matched_codes:
                active_rows.append(record)
                sheet_active_count += 1
                sheet_active_cells += numeric_cell_count
            else:
                record["quarantineReason"] = "fabric_code_not_in_fabric_code_list"
                quarantined_rows.append(record)
                sheet_quarantine_count += 1
                sheet_quarantine_cells += numeric_cell_count

        dimension_sheets.append(
            {
                "index": sheet_index,
                "name": sheet_name,
                "sourceRange": (
                    f"D1:{get_column_letter(numeric_columns[-1])}{max(data_rows)}"
                ),
                "columns": columns,
                "rawRowCount": sheet_raw_count,
                "activeRowCount": sheet_active_count,
                "quarantinedRowCount": sheet_quarantine_count,
                "rawNumericCellCount": sheet_raw_cells,
                "activeNumericCellCount": sheet_active_cells,
                "quarantinedNumericCellCount": sheet_quarantine_cells,
            }
        )

    return dimension_sheets, active_rows, quarantined_rows


def profile_definitions(dimension_sheets: list[dict]) -> list[dict]:
    definitions: list[dict] = []
    for dimension_sheet in dimension_sheets:
        contexts: OrderedDict[tuple, list[dict]] = OrderedDict()
        for column in dimension_sheet["columns"]:
            context = (
                column["orientation"],
                column["operatingSystem"],
                column["application"],
            )
            contexts.setdefault(context, []).append(column)

        for (orientation, operating_system, application), columns in contexts.items():
            common = [
                column
                for column in columns
                if column["tube"] is None or column["tube"].lower() == "all tubes"
            ]
            specific_tubes = list(
                OrderedDict.fromkeys(
                    column["tube"]
                    for column in columns
                    if column["tube"] is not None and column["tube"].lower() != "all tubes"
                )
            )
            tubes: list[str | None]
            if specific_tubes:
                tubes = specific_tubes
            elif any(column["tube"] and column["tube"].lower() == "all tubes" for column in common):
                tubes = ["all tubes"]
            else:
                tubes = [None]

            for tube in tubes:
                selected = list(common)
                if tube not in (None, "all tubes"):
                    selected.extend(column for column in columns if column["tube"] == tube)
                selected.sort(key=lambda column: column["sourceColumnIndex"])

                common_by_metric: dict[str, dict] = {}
                specific_by_metric: dict[str, dict] = {}
                common_duplicates: set[str] = set()
                specific_duplicates: set[str] = set()
                for column in selected:
                    metric = column["metric"]
                    if metric is None:
                        continue
                    target = (
                        common_by_metric
                        if column["tube"] is None or column["tube"].lower() == "all tubes"
                        else specific_by_metric
                    )
                    duplicates = common_duplicates if target is common_by_metric else specific_duplicates
                    if metric in target:
                        duplicates.add(metric)
                    target[metric] = column

                chosen = dict(common_by_metric)
                chosen.update(specific_by_metric)
                missing = [metric for metric in REQUIRED_LIMIT_METRICS if metric not in chosen]
                invalid_units = [
                    {
                        "metric": metric,
                        "sourceColumn": chosen[metric]["sourceColumn"],
                        "sourceUnit": chosen[metric]["sourceUnit"],
                    }
                    for metric in REQUIRED_LIMIT_METRICS
                    if metric in chosen and chosen[metric]["unit"] != "inch"
                ]
                ambiguous = sorted(
                    metric
                    for metric in common_duplicates | specific_duplicates
                    if metric in REQUIRED_LIMIT_METRICS
                )
                usable = not missing and not invalid_units and not ambiguous
                definition_id = f"roller-definition-{len(definitions) + 1:03d}"
                definitions.append(
                    {
                        "id": definition_id,
                        "sheet": dimension_sheet["name"],
                        "orientation": orientation,
                        "operatingSystem": operating_system,
                        "application": application,
                        "tube": tube,
                        "sourceColumns": [column["id"] for column in selected],
                        "sourceColumnsByMetric": {
                            metric: column["id"] for metric, column in chosen.items()
                        },
                        "sourceColumnLettersByMetric": {
                            metric: column["sourceColumn"] for metric, column in chosen.items()
                        },
                        "sourceUnitsByMetric": {
                            metric: column["sourceUnit"] for metric, column in chosen.items()
                        },
                        "missingRequiredMetrics": missing,
                        "invalidRequiredUnits": invalid_units,
                        "ambiguousRequiredMetrics": ambiguous,
                        "usable": usable,
                    }
                )
    return definitions


def build_profiles(
    dimension_sheets: list[dict],
    active_rows: list[dict],
    definitions: list[dict],
) -> tuple[list[dict], list[dict]]:
    columns_by_sheet = {
        sheet["name"]: {column["id"]: column for column in sheet["columns"]}
        for sheet in dimension_sheets
    }
    column_position_by_sheet = {
        sheet["name"]: {column["id"]: index for index, column in enumerate(sheet["columns"])}
        for sheet in dimension_sheets
    }
    usable_by_sheet: dict[str, list[dict]] = defaultdict(list)
    for definition in definitions:
        if definition["usable"]:
            usable_by_sheet[definition["sheet"]].append(definition)

    profiles: list[dict] = []
    profile_ids: dict[str, str] = {}
    assignments: list[dict] = []
    for row in active_rows:
        sheet_name = row["sheet"]
        for definition in usable_by_sheet[sheet_name]:
            limits: dict[str, int | float] = {}
            units: dict[str, str] = {}
            source_cells: dict[str, str] = {}
            for metric, column_id in definition["sourceColumnsByMetric"].items():
                column = columns_by_sheet[sheet_name][column_id]
                position = column_position_by_sheet[sheet_name][column_id]
                value = row["values"][position]
                if value is None:
                    raise SystemExit(
                        f"Missing numeric value for usable profile {definition['id']} at "
                        f"{sheet_name}!{column['sourceColumn']}{row['sourceRow']}"
                    )
                limits[metric] = value
                units[metric] = column["unit"]
                source_cells[metric] = f"{column['sourceColumn']}{row['sourceRow']}"

            key = json.dumps(
                {"definitionId": definition["id"], "limits": limits, "units": units},
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            )
            profile_id = profile_ids.get(key)
            if profile_id is None:
                profile_id = f"roller-profile-{len(profiles) + 1:04d}"
                profile_ids[key] = profile_id
                profiles.append(
                    {
                        "id": profile_id,
                        "definitionId": definition["id"],
                        "limits": limits,
                        "units": units,
                    }
                )

            assignments.append(
                {
                    "limitRowId": row["id"],
                    "profileDefinitionId": definition["id"],
                    "profileId": profile_id,
                    "fabricCodes": row["fabricCodes"],
                    "sourceCells": source_cells,
                }
            )
    return profiles, assignments


def region_counts(offerings: list[dict]) -> dict[str, int]:
    counts = {scope: 0 for scope in EXPECTED_REGION_COUNTS}
    for offering in offerings:
        counts[offering["regionScope"]] += 1
    return counts


def source_counts(
    offerings: list[dict],
    colors: list[dict],
    collections: list[str],
    dimension_sheets: list[dict],
    active_rows: list[dict],
    quarantined_rows: list[dict],
    definitions: list[dict],
    profiles: list[dict],
    assignments: list[dict],
) -> dict[str, int]:
    raw_rows = len(active_rows) + len(quarantined_rows)
    active_cells = sum(row["numericCellCount"] for row in active_rows)
    quarantine_cells = sum(row["numericCellCount"] for row in quarantined_rows)
    usable_definitions = sum(definition["usable"] for definition in definitions)
    return {
        "collections": len(collections),
        "colors": len(colors),
        "offerings": len(offerings),
        "fabricCodes": len({offering["fabricCode"] for offering in offerings}),
        "dimensionSheets": len(dimension_sheets),
        "dimensionColumns": sum(len(sheet["columns"]) for sheet in dimension_sheets),
        "rawLimitRows": raw_rows,
        "quarantinedLimitRows": len(quarantined_rows),
        "activeLimitRows": len(active_rows),
        "rawNumericCells": active_cells + quarantine_cells,
        "quarantinedNumericCells": quarantine_cells,
        "activeNumericCells": active_cells,
        "profileDefinitions": len(definitions),
        "usableProfileDefinitions": usable_definitions,
        "unusableProfileDefinitions": len(definitions) - usable_definitions,
        "limitProfiles": len(profiles),
        "profileAssignments": len(assignments),
    }


def validate_acceptance(
    ingested_hash: str,
    original_hash: str,
    data: dict[str, Any],
) -> None:
    if ingested_hash != EXPECTED_INGESTED_SHA256:
        raise SystemExit(
            f"Unexpected converted workbook SHA-256: {ingested_hash}; expected {EXPECTED_INGESTED_SHA256}"
        )
    if original_hash != EXPECTED_ORIGINAL_SHA256:
        raise SystemExit(
            f"Unexpected original workbook SHA-256: {original_hash}; expected {EXPECTED_ORIGINAL_SHA256}"
        )
    actual_counts = data["metadata"]["counts"]
    mismatches = {
        key: {"expected": expected, "actual": actual_counts.get(key)}
        for key, expected in EXPECTED_COUNTS.items()
        if actual_counts.get(key) != expected
    }
    if mismatches:
        raise SystemExit(f"Roller source count mismatch: {json.dumps(mismatches, sort_keys=True)}")
    if data["metadata"]["regionOfferingCounts"] != EXPECTED_REGION_COUNTS:
        raise SystemExit(
            "Regional offering counts do not match the authoritative Fabric Code List: "
            f"{data['metadata']['regionOfferingCounts']!r}"
        )
    if data["metadata"]["effectiveFrom"] != "2026-08-01":
        raise SystemExit(f"Unexpected effective date: {data['metadata']['effectiveFrom']}")
    quarantined = data["quarantinedLimitRows"]
    if not (
        len(quarantined) == 1
        and quarantined[0]["sourceFabricLabel"] == "AA0384"
        and quarantined[0]["numericCellCount"] == 40
    ):
        raise SystemExit("Expected only the 40-cell AA0384 orphan row to be quarantined")


TYPE_DECLARATIONS = """export interface NormanRollerV2ReleaseSourceRef {
  sheet: string;
  row: number;
  range: string;
}

export interface NormanRollerV2OfferingSourceRef extends NormanRollerV2ReleaseSourceRef {
  fabricCodeCell: string;
  collectionCell: string;
  colorCodeCell: string;
  colorNameCell: string;
  regionCell: string;
}

export interface NormanRollerV2Color {
  colorCode: string;
  colorName: string;
  collection: string;
  fabricCodes: string[];
  regionScopes: string[];
  effectiveFrom: string;
  offeringIds: string[];
}

export interface NormanRollerV2Offering {
  id: string;
  itemNumber: number | null;
  fabricCode: string;
  sourceFabricCode: string;
  collection: string;
  sourceCollection: string;
  colorCode: string;
  sourceColorCode: string;
  colorName: string;
  sourceColorName: string;
  regionScope: string;
  sourceRegionNote: string;
  effectiveFrom: string;
  sourceRef: NormanRollerV2OfferingSourceRef;
}

export interface NormanRollerV2HeaderCell {
  sourceCell: string;
  sourceRow: number;
  value: string;
}

export interface NormanRollerV2DimensionColumn {
  id: string;
  sourceColumn: string;
  sourceColumnIndex: number;
  headerPath: string[];
  sourceHeaderCells: NormanRollerV2HeaderCell[];
  orientation: string | null;
  operatingSystem: string | null;
  application: string | null;
  tube: string | null;
  metric: string | null;
  sourceMetric: string | null;
  unit: string | null;
  sourceUnit: string | null;
}

export interface NormanRollerV2DimensionSheet {
  index: number;
  name: string;
  sourceRange: string;
  columns: NormanRollerV2DimensionColumn[];
  rawRowCount: number;
  activeRowCount: number;
  quarantinedRowCount: number;
  rawNumericCellCount: number;
  activeNumericCellCount: number;
  quarantinedNumericCellCount: number;
}

export interface NormanRollerV2LimitRow {
  id: string;
  sheet: string;
  sourceRow: number;
  sourceCellRange: string;
  itemNumber: number | null;
  sourceItemCell: string;
  sourceFabricLabel: string;
  sourceFabricCell: string;
  sourceFabricTokens: string[];
  fabricCodes: string[];
  collection: string;
  sourceCollection: string;
  sourceCollectionCell: string;
  values: Array<number | null>;
  numericCellCount: number;
  quarantineReason?: string;
}

export interface NormanRollerV2InvalidUnit {
  metric: string;
  sourceColumn: string;
  sourceUnit: string | null;
}

export interface NormanRollerV2ProfileDefinition {
  id: string;
  sheet: string;
  orientation: string | null;
  operatingSystem: string | null;
  application: string | null;
  tube: string | null;
  sourceColumns: string[];
  sourceColumnsByMetric: Record<string, string>;
  sourceColumnLettersByMetric: Record<string, string>;
  sourceUnitsByMetric: Record<string, string | null>;
  missingRequiredMetrics: string[];
  invalidRequiredUnits: NormanRollerV2InvalidUnit[];
  ambiguousRequiredMetrics: string[];
  usable: boolean;
}

export interface NormanRollerV2LimitProfile {
  id: string;
  definitionId: string;
  limits: Record<string, number>;
  units: Record<string, string>;
}

export interface NormanRollerV2ProfileAssignment {
  limitRowId: string;
  profileDefinitionId: string;
  profileId: string;
  fabricCodes: string[];
  sourceCells: Record<string, string>;
}

export interface NormanRollerV2Counts {
  collections: number;
  colors: number;
  offerings: number;
  fabricCodes: number;
  dimensionSheets: number;
  dimensionColumns: number;
  rawLimitRows: number;
  quarantinedLimitRows: number;
  activeLimitRows: number;
  rawNumericCells: number;
  quarantinedNumericCells: number;
  activeNumericCells: number;
  profileDefinitions: number;
  usableProfileDefinitions: number;
  unusableProfileDefinitions: number;
  limitProfiles: number;
  profileAssignments: number;
}

export interface NormanRollerV2Metadata {
  revisionDate: string;
  effectiveFrom: string;
  releaseStatus: string;
  releaseStatusEvaluatedOn: string;
  activationPolicy: string;
  sourceSummary: string;
  sourceRef: NormanRollerV2ReleaseSourceRef;
  manufacturer: string;
  productFamily: string;
  sourceWorkbook: string;
  originalWorkbook: string;
  sourceWorkbookSha256: string;
  originalWorkbookSha256: string;
  inclusionRule: string;
  restrictionPolicy: string;
  regionOfferingCounts: Record<string, number>;
  counts: NormanRollerV2Counts;
}

export interface NormanRollerV2Source {
  metadata: NormanRollerV2Metadata;
  collections: string[];
  colors: NormanRollerV2Color[];
  offerings: NormanRollerV2Offering[];
  dimensionSheets: NormanRollerV2DimensionSheet[];
  limitRows: NormanRollerV2LimitRow[];
  quarantinedLimitRows: NormanRollerV2LimitRow[];
  profileDefinitions: NormanRollerV2ProfileDefinition[];
  limitProfiles: NormanRollerV2LimitProfile[];
  profileAssignments: NormanRollerV2ProfileAssignment[];
}
"""


def typed_chunks(
    base_name: str,
    type_name: str,
    values: list[dict],
    chunk_size: int = 250,
) -> tuple[str, list[str]]:
    declarations: list[str] = []
    references: list[str] = []
    for index in range(0, len(values), chunk_size):
        chunk_number = index // chunk_size + 1
        variable = f"{base_name}Chunk{chunk_number:03d}"
        payload = json.dumps(
            values[index : index + chunk_size],
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        )
        declarations.append(f"const {variable}: {type_name}[] = {payload};")
        references.append(variable)
    return "\n\n".join(declarations), references


def emit(output: Path, ingested_source: Path, original_source: Path, data: dict[str, Any]) -> None:
    ingested_hash = sha256(ingested_source)
    original_hash = sha256(original_source)
    validate_acceptance(ingested_hash, original_hash, data)
    source_base = dict(data)
    limit_profiles = source_base.pop("limitProfiles")
    profile_assignments = source_base.pop("profileAssignments")
    base_payload = json.dumps(source_base, ensure_ascii=False, indent=2, sort_keys=True)
    profile_declarations, profile_references = typed_chunks(
        "normanRollerV2LimitProfiles",
        "NormanRollerV2LimitProfile",
        limit_profiles,
    )
    assignment_declarations, assignment_references = typed_chunks(
        "normanRollerV2ProfileAssignments",
        "NormanRollerV2ProfileAssignment",
        profile_assignments,
    )
    profile_spreads = "\n".join(f"    ...{name}," for name in profile_references)
    assignment_spreads = "\n".join(f"    ...{name}," for name in assignment_references)
    content = (
        "// Generated by scripts/generate-norman-roller-v2-source.py. Do not edit.\n"
        f"// Converted source: {ingested_source.name}; SHA-256: {ingested_hash}\n"
        f"// Original source: {original_source.name}; SHA-256: {original_hash}\n\n"
        f"export const NORMAN_ROLLER_V2_INGESTED_WORKBOOK_SHA256 = {json.dumps(ingested_hash)} as const;\n"
        f"export const NORMAN_ROLLER_V2_ORIGINAL_WORKBOOK_SHA256 = {json.dumps(original_hash)} as const;\n"
        f"\n{TYPE_DECLARATIONS}\n"
        "const normanRollerV2SourceBase: Omit<NormanRollerV2Source, "
        f"\"limitProfiles\" | \"profileAssignments\"> = {base_payload};\n\n"
        f"{profile_declarations}\n\n"
        f"{assignment_declarations}\n\n"
        "export const normanRollerV2Source: NormanRollerV2Source = {\n"
        "  ...normanRollerV2SourceBase,\n"
        "  limitProfiles: [\n"
        f"{profile_spreads}\n"
        "  ],\n"
        "  profileAssignments: [\n"
        f"{assignment_spreads}\n"
        "  ],\n"
        "};\n"
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(content, encoding="utf-8")


def build_data(workbook: Any, ingested_source: Path, original_source: Path) -> dict[str, Any]:
    release = source_release(workbook)
    offerings, colors, collections = parse_fabric_catalog(workbook, release["effectiveFrom"])
    known_fabric_codes = {offering["fabricCode"] for offering in offerings}
    dimension_sheets, active_rows, quarantined_rows = parse_limit_sheets(
        workbook, known_fabric_codes
    )
    definitions = profile_definitions(dimension_sheets)
    profiles, assignments = build_profiles(
        dimension_sheets, active_rows, definitions
    )
    counts = source_counts(
        offerings,
        colors,
        collections,
        dimension_sheets,
        active_rows,
        quarantined_rows,
        definitions,
        profiles,
        assignments,
    )
    return {
        "metadata": {
            **release,
            "manufacturer": "Norman",
            "productFamily": "Roller Shades",
            "sourceWorkbook": ingested_source.name,
            "originalWorkbook": original_source.name,
            "sourceWorkbookSha256": sha256(ingested_source),
            "originalWorkbookSha256": sha256(original_source),
            "inclusionRule": (
                "A limit row is raw when a fabric label is physically authored in column B and at least "
                "one numeric dimension cell exists from column D onward; an item number in column A is not required. "
                "A raw row is active only when at least one exact fabric-code token resolves to Fabric Code List."
            ),
            "restrictionPolicy": (
                "Fail closed: create no selectable assignment when a required min/max width/height "
                "dimension is absent, ambiguous, or not expressed in inches."
            ),
            "regionOfferingCounts": region_counts(offerings),
            "counts": counts,
        },
        "collections": collections,
        "colors": colors,
        "offerings": offerings,
        "dimensionSheets": dimension_sheets,
        "limitRows": active_rows,
        "quarantinedLimitRows": quarantined_rows,
        "profileDefinitions": definitions,
        "limitProfiles": profiles,
        "profileAssignments": assignments,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path, help="Converted .xlsx source workbook")
    parser.add_argument("output", type=Path, help="Generated TypeScript output")
    parser.add_argument(
        "--original-workbook",
        type=Path,
        help="Original .xls artifact; inferred from a sibling converted/ directory when omitted",
    )
    args = parser.parse_args()
    original_source = source_original_for(args.workbook, args.original_workbook)
    workbook = load_workbook(args.workbook, data_only=True, read_only=False)
    data = build_data(workbook, args.workbook, original_source)
    emit(args.output, args.workbook, original_source, data)


if __name__ == "__main__":
    main()
